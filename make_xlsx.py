#!/usr/bin/env python3
"""Convert the ranked submission CSV to XLSX (the format the upload form expects).

Usage:
    python make_xlsx.py --in submission.csv --out submission.xlsx

Preserves the exact column order produced by the ranker
(candidate_id, rank, score, reasoning) so the organizer's checks still pass.
"""
from __future__ import annotations
import argparse, csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--in", dest="src", default="submission/sample_output.csv")
    p.add_argument("--out", dest="dst", default="submission.xlsx")
    a = p.parse_args()
    rows = list(csv.reader(Path(a.src).open(encoding="utf-8")))
    if not rows:
        raise SystemExit("empty input")
    wb = Workbook(); ws = wb.active; ws.title = "ranking"
    for r in rows:
        ws.append(r)
    # header styling + keep score as text so 6-dp precision is never reformatted
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 90
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=False, vertical="top")
    wb.save(a.dst)
    print(f"wrote {ws.max_row-1} ranked rows -> {a.dst}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())